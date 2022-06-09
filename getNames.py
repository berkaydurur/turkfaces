import requests
from openpyxl import Workbook
from lxml import html
import time
wb = Workbook()

def parseNames():
    i= 0
    jobArr = []
    nameArr = []
    x=0
    for t in range(20):

        response = requests.get('https://www.imdb.com/search/name/?birth_date=1900-01-01,2022-12-31&birth_place=Turkey&gender=male,female&adult=include&count=250&start='+str(x+1)+'&ref_=rlm')
        print('https://www.imdb.com/search/name/?birth_date=1900-01-01,2022-12-31&birth_place=Turkey&gender=male,female&adult=include&count=250&start='+str(x+1)+'&ref_=rlm')
        tree = html.fromstring(response.content)
        jobs = tree.xpath('//div[@class="lister-item-content"]/p[@class="text-muted text-small"]/text()')
        names = tree.xpath('//h3[@class="lister-item-header"]/a/text()')
        nameArr = nameArr + names
        jobArr = jobArr + jobs
        
        deneme = []
        x+=250
    getNames(nameArr, jobArr)

def getNames(nameArr, jobArr):
    getJobs = []
    getNames = []  
    ws = wb.active
    ws.title = "kisiler"
    k= 0 
    for i in range(len(jobArr)):
        if(jobArr[i].strip() != "\n" or jobArr[i].strip() != ""):
            if(jobArr[i].strip() == "Actor" or jobArr[i].strip() == "Actress"):
                getJobs.append(jobArr[i].strip(" "))
                getNames.append(nameArr[k].strip(" "))
                ws.cell(row=k+1, column=1, value=nameArr[k])
                ws.cell(row=k+1, column=2, value=jobArr[i])
                k = k+1
    #print(getJobs)
    print(len(jobArr), len(nameArr))
    print(len(getJobs), len(getNames))

    file_object = open('turkish_actors_and_actresses.txt','w')
    wb.save('names.xlsx')
    for name in getNames : 
        file_object.write(name)
    file_object.close()

parseNames()